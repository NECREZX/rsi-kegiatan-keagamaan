#!/usr/bin/env python3
import sys
import json
import openpyxl

MONTHS_MAP = {
    '01': 'Jan', '02': 'Feb', '03': 'Mar', '04': 'Apr',
    '05': 'Mei', '06': 'Jun', '07': 'Jul', '08': 'Agust',
    '09': 'Sep', '10': 'Okt', '11': 'Nop', '12': 'Des'
}

def update_xlsx(xlsx_file, json_file):
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    wb = openpyxl.load_workbook(xlsx_file)
    ws = wb['REKAP KEGIATAN KEAGAMAAN'] if 'REKAP KEGIATAN KEAGAMAAN' in wb.sheetnames else wb.active
    
    # 1. Map Headers (Row 5)
    header_row = 5
    col_map = {} # {'N A M A': 2, ...}
    
    # Standard profile columns
    PROFILE_COLS = {
        'nama': ['N A M A', 'NAMA'],
        'nik': ['N I K', 'N I K ', 'NIK'],
        'status_pegawai': ['STATUS PEGAWAI'],
        'jk': ['JK', 'J/K', 'JENIS KELAMIN'],
        'kelompok_nakes': ['KELOMPOK NAKES'],
        'nama_jabatan': ['NAMA JABATAN', 'JABATAN'],
        'struktur_lini': ['STRUKTUR LINI'],
        'tempat_tugas': ['TEMPAT TUGAS'],
        'no': ['NO', 'NO.'],
        'tanggal_berhenti': ['TANGGAL BERHENTI', 'TGL BERHENTI']
    }
    
    # Pre-scan headers
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val:
            s_val = str(val).strip().upper()
            col_map[s_val] = col
            
    # Ensure TANGGAL BERHENTI column exists
    if not any(k in col_map for k in ['TANGGAL BERHENTI', 'TGL BERHENTI']):
        new_col = ws.max_column + 1
        ws.cell(row=header_row, column=new_col).value = 'TANGGAL BERHENTI'
        col_map['TANGGAL BERHENTI'] = new_col
            
    # Helper to find col index for a profile field
    def get_col_idx(field):
        candidates = PROFILE_COLS.get(field, [])
        for c in candidates:
            if c in col_map: return col_map[c]
            # Try partial match if needed? No, exact is safer
        return None

    # 2. Map existing rows by NO (or NIK/Nama if needed, but NO is safest ID)
    # Assumes NO in Excel == id in JSON
    row_map = {} # {id: row_num}
    no_col = get_col_idx('no')
    
    if no_col:
        for r in range(header_row + 1, ws.max_row + 1):
            val = ws.cell(row=r, column=no_col).value
            if val and str(val).strip().isdigit():
                row_map[int(val)] = r
    
    # 3. Update or Add rows
    # Track which Excel rows are active (found in JSON)
    active_rows = set()
    
    for peg in data['pegawai']:
        pid = peg.get('id')
        if not pid: continue
        
        row_num = row_map.get(pid)
        
        # If new employee, append to bottom
        if not row_num:
            row_num = ws.max_row + 1
            ws.cell(row=row_num, column=no_col).value = pid
            row_map[pid] = row_num # Mark as mapped
            
        active_rows.add(row_num)

        # Write Profile Data
        for field in PROFILE_COLS:
            if field == 'no': continue
            c_idx = get_col_idx(field)
            if c_idx:
                val = peg.get(field, '')
                ws.cell(row=row_num, column=c_idx).value = val

        # Write Activity Data
        for year, months in peg.get('data', {}).items():
            yr_short = year[-2:]
            for month, vals in months.items():
                mon_str = MONTHS_MAP.get(month, '')
                if not mon_str: continue
                
                # Activity columns mapping
                # Columns like "Mengaji Jan-26"
                act_cols = {
                    'mengaji': f'Mengaji {mon_str}-{yr_short}',
                    'kajian_fiqih': f'Kajian Fiqih {mon_str}-{yr_short}',
                    'phbi': f'PHBI {mon_str}-{yr_short}'
                }
                
                for act_key, col_name in act_cols.items():
                    # Find column for this specific activity-month-year
                    # Check case-insensitive match against col_map keys
                    c_idx = None
                    col_name_upper = col_name.upper()
                    
                    # Direct lookup
                    if col_name_upper in col_map:
                        c_idx = col_map[col_name_upper]
                    else:
                        # Fuzzy lookup (spaces etc)
                        for k, v in col_map.items():
                            if k.replace(' ', '') == col_name_upper.replace(' ', ''):
                                c_idx = v
                                break
                    
                    if c_idx:
                        v = int(vals.get(act_key, 0))
                        ws.cell(row=row_num, column=c_idx).value = v if v > 0 else None
    
    # 4. Delete rows not in JSON (Bottom-up to avoid shifting issues)
    # Get all rows that were mapped initially (from step 2)
    # Identify which ones are NOT in active_rows
    rows_to_delete = []
    for pid, r_num in row_map.items():
        if r_num not in active_rows:
            rows_to_delete.append(r_num)
            
    # Sort descending
    rows_to_delete.sort(reverse=True)
    
    for r in rows_to_delete:
        ws.delete_rows(r)
        print(f"Deleted row {r}")

    wb.save(xlsx_file)
    print(f"Updated {xlsx_file}")

if __name__ == '__main__':
    update_xlsx(sys.argv[1], sys.argv[2])
